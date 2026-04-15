"""
fill_crm.py — Điền thông tin đơn hàng vào mau crm.xlsx
Usage:
    python fill_crm.py '<json_data>' [input_file] [output_file]

JSON format:
{
  "khach_hang": "Tên khách hàng",
  "ma_so_thue": "0101234567",
  "dia_chi_hoa_don": "123 Nguyễn Trãi, Hà Nội",
  "ngay_dat_hang": "11/04/2026",
  "so_don_hang": "DH-20260411-001",
  "nguoi_nhan": "Nguyễn Văn B",
  "dia_chi_giao_hang": "456 Lê Lợi, TP.HCM",
  "sdt_nguoi_nhan": "0987654321",
  "chiet_khau": 0,
  "thue_gtgt": 0,
  "san_pham": [
    {"ten": "Laptop Dell Inspiron 15", "dvt": "Cái", "sl": 1, "don_gia": 15000000},
    {"ten": "Chuột không dây Logitech", "dvt": "Cái", "sl": 2, "don_gia": 300000}
  ]
}
"""

import sys
import json
import shutil
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from copy import copy

sys.stdout.reconfigure(encoding="utf-8")

INSERT_AT = 14          # Dòng đầu tiên chứa sản phẩm trong template
SUMMARY_BASE = 16       # Dòng "Tổng thành tiền" trong template gốc
PAYMENT_BASE = 19       # Dòng "Tổng tiền thanh toán" trong template gốc


def thin_border():
    s = Side(border_style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def copy_row_style(ws, source_row, target_row, max_col=7):
    """Copy font, fill, alignment, border từ source_row sang target_row."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)
        if src.font:
            tgt.font = copy(src.font)
        if src.fill:
            tgt.fill = copy(src.fill)
        if src.alignment:
            tgt.alignment = copy(src.alignment)
        if src.border:
            tgt.border = copy(src.border)
        if src.number_format:
            tgt.number_format = src.number_format


def fill_crm(data: dict, input_file: str, output_file: str):
    wb = load_workbook(input_file)
    ws = wb.active

    products = data.get("san_pham", [])
    n = len(products)

    # ================================================================
    # BƯỚC 1: Lưu và unmerge tất cả merged cells từ INSERT_AT trở đi
    # ================================================================
    merges_to_move = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= INSERT_AT:
            merges_to_move.append(str(mr))
            ws.unmerge_cells(str(mr))

    # ================================================================
    # BƯỚC 2: Insert N dòng mới tại INSERT_AT
    # ================================================================
    ws.insert_rows(INSERT_AT, amount=n)

    # ================================================================
    # BƯỚC 3: Re-merge cells tại vị trí mới (dịch xuống n dòng)
    # ================================================================
    for merge_str in merges_to_move:
        min_col, min_row, max_col_val, max_row = range_boundaries(merge_str)
        new_range = (
            f"{get_column_letter(min_col)}{min_row + n}"
            f":{get_column_letter(max_col_val)}{max_row + n}"
        )
        ws.merge_cells(new_range)

    # ================================================================
    # BƯỚC 4: Điền dữ liệu sản phẩm + copy style từ dòng tiêu đề
    # ================================================================
    border = thin_border()
    currency_fmt = '#,##0'
    for i, product in enumerate(products):
        row = INSERT_AT + i
        copy_row_style(ws, 13, row)          # copy style từ header row
        ws.cell(row=row, column=1, value=i + 1)                          # STT
        ws.cell(row=row, column=3, value=product.get("ten", ""))         # Tên SP
        ws.cell(row=row, column=4, value=product.get("dvt", "Cái"))      # ĐVT
        ws.cell(row=row, column=5, value=product.get("sl", 0))            # SL
        ws.cell(row=row, column=6, value=product.get("don_gia", 0))       # Đơn giá
        ws.cell(row=row, column=7).value = f"=E{row}*F{row}"              # Thành tiền

        # Đảm bảo viền và format tiền tệ
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col in (5, 6, 7):
                cell.number_format = currency_fmt

    # ================================================================
    # BƯỚC 5: Cập nhật công thức tổng hợp
    # ================================================================
    summary_row = SUMMARY_BASE + n      # Dòng "Tổng thành tiền"
    discount_row = summary_row + 1      # Dòng "Tiền chiết khấu"
    tax_row = summary_row + 2           # Dòng "Thuế GTGT"
    payment_row = PAYMENT_BASE + n      # Dòng "Tổng tiền thanh toán"
    last_product_row = INSERT_AT + n - 1

    ws.cell(row=summary_row, column=6).value = f"=SUM(G{INSERT_AT}:G{last_product_row})"
    ws.cell(row=payment_row, column=6).value = f"=F{summary_row}-F{discount_row}-F{tax_row}"

    # Điền chiết khấu và thuế nếu có
    if data.get("chiet_khau", 0):
        ws.cell(row=discount_row, column=6, value=data["chiet_khau"])
    if data.get("thue_gtgt", 0):
        ws.cell(row=tax_row, column=6, value=data["thue_gtgt"])

    # ================================================================
    # BƯỚC 6: Điền thông tin header đơn hàng
    # ================================================================
    def set_val(cell_addr, prefix, value):
        ws[cell_addr] = f"{prefix}{value}" if prefix else value

    if data.get("khach_hang"):
        ws["A4"] = f"Khách hàng: {data['khach_hang']}"
    if data.get("ngay_dat_hang"):
        ws["E4"] = f"Ngày đặt hàng: {data['ngay_dat_hang']}"
    if data.get("ma_so_thue"):
        ws["A5"] = f"Mã số thuế: {data['ma_so_thue']}"
    if data.get("so_don_hang"):
        ws["E5"] = f"Số đơn hàng: {data['so_don_hang']}"
    if data.get("dia_chi_hoa_don"):
        ws["A6"] = f"Địa chỉ hóa đơn: {data['dia_chi_hoa_don']}"
    if data.get("nguoi_nhan"):
        ws["A9"] = f"Người nhận hàng: {data['nguoi_nhan']}"
    if data.get("sdt_nguoi_nhan"):
        ws["E9"] = f"SĐT người nhận: {data['sdt_nguoi_nhan']}"
    if data.get("dia_chi_giao_hang"):
        ws["A10"] = f"Địa chỉ giao hàng: {data['dia_chi_giao_hang']}"

    wb.save(output_file)
    print(f"✅ Đã lưu file: {output_file}")


if __name__ == "__main__":
    # Usage A: python fill_crm.py --json-file data.json [input.xlsx] [output.xlsx]
    # Usage B: python fill_crm.py '{...json...}' [input.xlsx] [output.xlsx]
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python fill_crm.py --json-file data.json [input.xlsx] [output.xlsx]")
        print("  python fill_crm.py '<json_string>' [input.xlsx] [output.xlsx]")
        sys.exit(1)

    if sys.argv[1] == "--json-file":
        json_file = sys.argv[2]
        with open(json_file, encoding="utf-8-sig") as f:
            order_data = json.load(f)
        input_xlsx  = sys.argv[3] if len(sys.argv) > 3 else "mau crm.xlsx"
        output_xlsx = sys.argv[4] if len(sys.argv) > 4 else "don_hang_output.xlsx"
    else:
        raw_json = sys.argv[1]
        input_xlsx  = sys.argv[2] if len(sys.argv) > 2 else "mau crm.xlsx"
        output_xlsx = sys.argv[3] if len(sys.argv) > 3 else "don_hang_output.xlsx"
        order_data = json.loads(raw_json)

    fill_crm(order_data, input_xlsx, output_xlsx)
